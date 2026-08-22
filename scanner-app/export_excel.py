import json
import sys
from copy import copy
from datetime import datetime

from openpyxl import load_workbook


def main(template_path, data_path, output_path):
    with open(data_path, "r", encoding="utf-8") as source:
        records = json.load(source)

    workbook = load_workbook(template_path)
    sheet = workbook["Hoja1"]
    min_column, max_column = 1, 5
    max_row = sheet.max_row
    style_row = max_row
    for record in records:
        target_row = max_row + 1
        for column in range(min_column, max_column + 1):
            source_cell = sheet.cell(style_row, column)
            target_cell = sheet.cell(target_row, column)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)

        sheet.cell(target_row, min_column).value = datetime.strptime(record["date"], "%Y-%m-%d")
        sheet.cell(target_row, min_column + 1).value = record["amount"]
        sheet.cell(target_row, min_column + 2).value = record["cuit"]
        sheet.cell(target_row, min_column + 3).value = record["invoice"]
        sheet.cell(target_row, min_column + 4).value = record["concept"]
        max_row = target_row

    workbook.save(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 4:
        raise SystemExit("Uso: export_excel.py <modelo.xlsx> <gastos.json> <salida.xlsx>")
    main(*sys.argv[1:])
