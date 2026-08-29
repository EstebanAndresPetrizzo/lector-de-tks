from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, Mapping, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = ["FECHA", "IMPORTE", "CUIT sin guiones", "NRO FACTURA", "CONCEPTO del GATOS"]


def build_workbook(records: Iterable[Mapping[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hoja1"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    sheet.page_margins = PageMargins(left=0.25, right=0.25, top=0.45, bottom=0.45, header=0.15, footer=0.15)

    header_fill = PatternFill("solid", fgColor="BFBFBF")
    header_font = Font(name="Calibri", size=11, bold=True, italic=True, color="000000")
    data_font = Font(name="Calibri", size=11, color="FF0000")
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=medium if column == 1 else thin,
            right=medium if column == len(HEADERS) else thin,
            top=medium,
            bottom=medium,
        )

    rows = list(records)
    for row_number, record in enumerate(rows, start=2):
        date_value = datetime.strptime(str(record["date"]), "%Y-%m-%d")
        values = [
            date_value,
            round(float(record["amount"]), 2),
            str(record["cuit"]),
            str(record["invoice"]),
            str(record["concept"]),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sheet.cell(row_number, 1).number_format = "dd/mm/yy"
        sheet.cell(row_number, 2).number_format = '#,##0.00'
        sheet.cell(row_number, 3).number_format = "@"
        invoice = str(record["invoice"])
        significant_digits = invoice.lstrip("0") or "0"
        if invoice.isdigit() and len(significant_digits) <= 15:
            sheet.cell(row_number, 4).value = int(invoice)
            sheet.cell(row_number, 4).number_format = "0" * len(invoice)
        else:
            sheet.cell(row_number, 4).number_format = "@"
            sheet.cell(row_number, 4).quotePrefix = True

    sheet.row_dimensions[1].height = 42
    widths = {"A": 15, "B": 18, "C": 23, "D": 23, "E": 38}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    if rows:
        sheet.print_area = f"A1:E{len(rows) + 1}"
        table = Table(displayName="RegistroDeducciones", ref=f"A1:E{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
