from __future__ import annotations

import base64
import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from lector_tks.excel_export import HEADERS, build_workbook
from lector_tks.extractor import normalize_invoice, parse_fiscal_qr, parse_ocr_text
from lector_tks.storage import RecordStore

# Todos los CUIT, comprobantes e importes de este archivo son ejemplos ficticios.

class ExtractorTests(unittest.TestCase):
    def test_parses_arca_qr(self):
        data = {
            "fecha": "2025-12-05",
            "cuit": "00000000000",
            "ptoVta": 17,
            "nroCmp": 123456,
            "importe": 1234.56,
        }
        payload = base64.urlsafe_b64encode(json.dumps(data).encode()).decode().rstrip("=")
        result = parse_fiscal_qr(f"https://www.afip.gob.ar/fe/qr/?p={payload}")
        self.assertIsNotNone(result)
        self.assertEqual(result["date"], "2025-12-05")
        self.assertEqual(result["cuit"], "00000000000")
        self.assertEqual(result["invoice"], "0001700123456")
        self.assertEqual(result["amount"], 1234.56)

    def test_parses_common_ocr_layout(self):
        text = """
        FARMACIA EJEMPLO
        CUIT: 00-00000000-0
        Fecha de emisión: 05/11/2025
        Factura B 00007-00001234
        TOTAL $ 1.234,56
        """
        result = parse_ocr_text(text, "GASTOS DE RESPRESENTACION")
        self.assertEqual(result["date"], "2025-11-05")
        self.assertEqual(result["amount"], 1234.56)
        self.assertEqual(result["cuit"], "00000000000")
        self.assertEqual(result["invoice"], "0000700001234")
        self.assertEqual(result["concept"], "FARMACIA")
        self.assertEqual(result["status"], "ready")

    def test_invoice_keeps_leading_zeroes(self):
        self.assertEqual(normalize_invoice("7", "1234"), "0000700001234")
        self.assertEqual(normalize_invoice("0045", "000000123456"), "0045000000123456")

    def test_parses_abbreviated_point_of_sale(self):
        text = "P. de Venta 0045 Nro 000000123456\nFecha: 07-07-2026\nCuit: 00000000000\nTotal: $ 12.345,67"
        result = parse_ocr_text(text, "GASTOS DE RESPRESENTACION")
        self.assertEqual(result["invoice"], "0045000000123456")

    def test_parses_amount_with_ocr_spacing_before_cents(self):
        text = "TOTAL 12345 , 67"
        result = parse_ocr_text(text, "GASTOS DE RESPRESENTACION")
        self.assertEqual(result["amount"], 12345.67)


class StorageTests(unittest.TestCase):
    def test_duplicate_is_not_added_twice(self):
        with tempfile.TemporaryDirectory() as directory:
            store = RecordStore(Path(directory) / "records.sqlite3")
            record = {
                "date": "2025-11-05", "amount": 1234.56, "cuit": "00000000000",
                "invoice": "0000700001234", "concept": "FARMACIA", "confidence": 100,
                "warnings": [], "status": "ready",
            }
            first, first_duplicate = store.add_record(record)
            second, second_duplicate = store.add_record(record)
            self.assertFalse(first_duplicate)
            self.assertTrue(second_duplicate)
            self.assertEqual(first["id"], second["id"])
            self.assertEqual(len(store.list_records()), 1)


class ExcelTests(unittest.TestCase):
    def test_workbook_matches_required_columns_and_types(self):
        records = [
            {
                "date": "2025-11-05", "amount": 1234.56, "cuit": "00000000000",
                "invoice": "0000700001234", "concept": "FARMACIA",
            },
            {
                "date": "2026-07-07", "amount": 9876.54, "cuit": "00000000000",
                "invoice": "0045000000123456", "concept": "GASTOS DE RESPRESENTACION",
            },
        ]
        content = build_workbook(records)
        workbook = load_workbook(BytesIO(content))
        sheet = workbook["Hoja1"]
        self.assertEqual([sheet.cell(1, column).value for column in range(1, 6)], HEADERS)
        self.assertEqual(sheet["C2"].value, "00000000000")
        self.assertEqual(sheet["D2"].value, 700001234)
        self.assertEqual(sheet["C2"].number_format, "@")
        self.assertEqual(sheet["D2"].number_format, "0000000000000")
        self.assertEqual(sheet["B2"].value, 1234.56)
        self.assertEqual(sheet["D3"].value, 45000000123456)
        self.assertEqual(sheet["D3"].number_format, "0000000000000000")


if __name__ == "__main__":
    unittest.main()
